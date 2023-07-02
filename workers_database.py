import tkinter.ttk
from tkinter import *
from PIL import ImageTk, Image
import sqlite3
import csv
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
from tkinter import filedialog
from tkinter import ttk

if __name__ == "__main__":
    root = Tk()
    root.title("Workers Database")
    root.geometry("725x705")
    root.config(background="#DBF0D7")
    icon = PhotoImage(file="workers.logo.png")
    root.iconphoto(True, icon)

    conn = sqlite3.connect("workers_cost_file.db")
    c = conn.cursor()


    c.execute("""CREATE TABLE IF NOT EXISTS  workers (
                date integer,
                name text,
                work text,
                hours integer,
                price float)""")

    conn.commit()


    #worker_combobox = None

    worker_combobox = ttk.Combobox(root, font=("arial black", 13), width=15)
    worker_combobox.grid(row=7, columnspan=2,  pady=10)



    def update_worker_list():
        global worker_combobox
        c.execute("SELECT DISTINCT name FROM workers")
        workers = [row[0] for row in c.fetchall()]
        worker_combobox['values'] = workers




    title_label = Label(root, text="     ΕΡΓΑΤΕΣ     ", font=("arial black", 20), background="black",
                        relief="ridge", borderwidth=7, foreground="yellow")
    title_label.grid(row=0, columnspan=2, padx=210, pady=15)


    def submit():
        conn = sqlite3.connect("workers_cost_file.db")
        c = conn.cursor()

        c.execute("INSERT INTO workers VALUES (:date, :name, :work, :hours, :price)",
                  {
                      "date": date.get(),
                      "name": name.get(),
                      "work": work.get(),
                      "hours": hours.get(),
                      "price": price.get()
                  })
        price_value = price.get()
        if price_value == "0" or not price_value.strip():
            price_value = ""

        confirmation = messagebox.askquestion("Εισαγωγή Αρχείων", "Do You Want To Insert this Files ???")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Εισαγωγή Αρχείων ", "Επιτυχής Εισαγωγή Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η Εισαγωγή Αρχείων ακυρώθηκε. !!!")
        conn.close()

        date.delete(0, END)
        name.delete(0, END)
        work.delete(0, END)
        hours.delete(0, END)
        price.delete(0, END)

    def save_to_excel():
        conn = sqlite3.connect("workers_cost_file.db")
        c = conn.cursor()

        c.execute("SELECT * FROM workers")
        data = c.fetchall()

        try:
            workbook = load_workbook("total.workers.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            headers = ["ΗΜΕΡΟΜΗΝΙΑ", "ΟΜΟΜΑ", "ΕΡΓΑΣΙΑ", "ΩΡΕΣ", "ΤΙΜΗ"]
            sheet.append(headers)

        # Clear existing data in the worksheet
        sheet.delete_rows(2, sheet.max_row)

        for row in data:
            sheet.append(row)

        workbook.save("total.workers.xlsx")

        conn.close()

        messagebox.showinfo("Αποθήκευση Δεδομένων", "Τα δεδομένα αποθηκεύτηκαν σε αρχείο Excel !!!")


    def delete_all_data():
        confirmation = messagebox.askyesno("Confirmation", "Είστε σίγουρος ότι θέλετε να διαγράψετε όλα τα Αρχεία ;")
        if confirmation:
            conn = sqlite3.connect("workers_cost_file.db")
            c = conn.cursor()
            c.execute("DELETE FROM workers")
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Ολα τα Αρχεία διαγράφηκαν με επιτυχία !!")



    def show():
        show = Tk()
        show.title("WORKERS DATABASE")
        show.geometry("600x650")
        show.config(background="#DBF0D7")



        show_id = choose_id.get()
        conn = sqlite3.connect("workers_cost_file.db")
        c = conn.cursor()

        c.execute("SELECT *,  oid FROM workers")
        elements = c.fetchall()

        for i, element in enumerate(elements):
            text = '   '.join(str(item) for item in element)
            show_label = Label(show, text=text, font=("arial black", 10), relief="ridge", bd=7, borderwidth=5,
                               background="#D1F3F3", foreground="#20047A")
            show_label.grid(row=i, column=0, pady=1, sticky=W, padx=20, ipadx=30)

        save_button = Button(show, text="Αποθήκευση σε 'Excel' ", command=save_to_excel, font=("arial black", 11),
                             background="#06D784")
        save_button.grid(row=i + 1, column=0, pady=10, ipadx=26, sticky=W, padx=20)
        delete_button = Button(show, text="Διαγραφή όλων των Αρχείων", command=delete_all_data, font=("arial black", 11),
                               background="#EA6969")
        delete_button.grid(row=i + 1, column=0, pady=10, padx=320, ipadx=8)






    def update():
        conn = sqlite3.connect("workers_cost_file.db")
        c = conn.cursor()
        show_id = choose_id.get()
        c.execute("""UPDATE workers SET
               date = :date,
               name = :name,
               work = :work,
               hours = :hours,
               price = :price
    
    
               WHERE oid = :oid""",
                  {'date': date_editor.get(),
                   'name': name_editor.get(),
                   'work': work_editor.get(),
                   'hours': hours_editor.get(),
                   'price': price_editor.get(),
                   'oid': show_id})

        confirmation = messagebox.askquestion("Επεξεργασία Αρχείων", "Do you want to save the changes  ???")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Αποθήκευση Αρχείων ", "Επιτυχής αλλαγή και αποθήκευση Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η αλλαγή Αρχείων ακυρώθηκε. !!!")

        conn.close()
        editor.destroy()

    def edit():
        global editor
        editor = Tk()
        editor.title("ΕΠΕΞΕΡΓΑΣΙΑ ΣΤΟΙΧΕΙΩΝ")
        editor.geometry("500x400")
        editor.config(background="#DBF0D7")


        conn = sqlite3.connect("workers_cost_file.db")
        c = conn.cursor()

        show_id = choose_id.get()

        c.execute("SELECT * FROM workers WHERE oid = " + show_id)
        shows = c.fetchall()

        global date_editor
        global name_editor
        global work_editor
        global hours_editor
        global price_editor

        date_editor_label = Label(editor, text=" Ημερομηνία : ", font=("arial black", 12), background="#DBF0D7",
                           foreground="#110464")
        date_editor_label.grid(row=1, column=0, sticky=W, padx=20, pady=(20, 10))
        name_editor_label = Label(editor, text=" Όνομα : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
        name_editor_label.grid(row=2, column=0, sticky=W, padx=20, pady=10)
        work_editor_label = Label(editor, text=" Εργασία : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
        work_editor_label.grid(row=3, column=0, sticky=W, padx=20, pady=10)
        hours_editor_label = Label(editor, text=" Ώρες : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
        hours_editor_label.grid(row=4, column=0, sticky=W, padx=20, pady=10)
        price_editor_label = Label(editor, text=" Ημερομίσθιο € : ", font=("arial black", 12), background="#DBF0D7",
                            foreground="#110464")
        price_editor_label.grid(row=5, column=0, sticky=W, padx=20, pady=10)

        date_editor = Entry(editor, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
        date_editor.grid(row=1, column=1, padx=20, sticky=W, pady=(20, 10))
        name_editor = Entry(editor, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
        name_editor.grid(row=2, column=1, padx=20, sticky=W, pady=10)
        work_editor = Entry(editor, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
        work_editor.grid(row=3, column=1, padx=20, sticky=W, pady=10)
        hours_editor = Entry(editor, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
        hours_editor.grid(row=4, column=1, padx=20, sticky=W, pady=10)
        price_editor = Entry(editor, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
        price_editor.grid(row=5, column=1, padx=20, sticky=W, pady=10)

        for show in shows:
            date_editor.insert(0, show[0])
            name_editor.insert(0, show[1])
            work_editor.insert(0, show[2])
            hours_editor.insert(0, show[3])
            price_editor.insert(0, show[4])


        edit_btn = Button(editor, text="Επεξεργασία και Αποθήκευση \n Στοιχείων", bd=4, command=update, font=("arial black", 10), background="#06D784", activeforeground="#06D784", activebackground="#06D784")
        edit_btn.grid(row=7, column=0, columnspan=2, pady=30, padx=10, ipadx=30)

    def delete():
        conn = sqlite3.connect("workers_cost_file.db")
        c = conn.cursor()

        c.execute("DELETE from workers WHERE oid = " + choose_id.get())

        confirmation = messagebox.askquestion("Διαγραφή Στοιχείων", " Θέλετε να διαγράψετε τα Στοιχεία ;")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo("Η Διαγραφή Ολοκληρώθηκε", "Τα στοιχεία διαγράφηκαν με επιτυχία. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Διαγραφής", "Η διαγραφή ακυρώθηκε. !!!")

        conn.commit()
        conn.close()



    def select_worker():
        selected_worker = worker_combobox.get()
        if selected_worker:
            c.execute("SELECT name, SUM(price), SUM(hours), work  FROM workers WHERE name=? GROUP BY name", (selected_worker,))
            worker_info = c.fetchall()
            if worker_info:
                worker_window = Toplevel(root)
                worker_window.title("Worker Information")
                worker_window.geometry("550x300")
                worker_window.config(background="#DBF0D7")

                info_labels = [" ΌΝΟΜΑ : ", " ΣΥΝΟΛΟ ΕΥΡΩ : ", " ΣΥΝΟΛΙΚΕΣ ΩΡΕΣ : ", " ΕΡΓΑΣΙΑ : "]

                for row_index, worker_data in enumerate(worker_info):
                    for i, label in enumerate(info_labels):
                        label_text = Label(worker_window, text=label, font=("arial black", 12), width=20, relief="ridge", bd=3,  background="#F1B52D")
                        label_text.grid(row=row_index * len(info_labels) + i, column=0, padx=20, sticky=W, pady=10)

                        if i == 1:
                            # Display the sum of prices
                            value_label = Label(worker_window, text=worker_data[i], font=("arial black", 13), foreground="#62057C",  background="#DBF0D7")
                        else:
                            # Display the worker name
                            value_label = Label(worker_window, text=worker_data[i], font=("arial black", 13), foreground="#62057C",  background="#DBF0D7")

                        value_label.grid(row=row_index * len(info_labels) + i, column=1, padx=20, pady=10)

            else:
                messagebox.showerror("Error", "No information available for the selected worker.")
        else:
            messagebox.showwarning("Warning", "Please select a worker.")

        update_worker_list()



    update_worker_list()

    def open_cost_worker():
        total_w = Tk()
        total_w.title("ΣΥΝΟΛΙΚΟ ΚΟΣΤΟΣ ΕΡΓΑΤΩΝ")
        total_w.geometry("650x350")
        total_w.config(background="#DBF0D7")


        def total_cost_workers():
            conn = sqlite3.connect("workers_cost_file.db")
            c = conn.cursor()

            c.execute("SELECT price FROM workers")
            rows = c.fetchall()
            total_euros = sum(
                float(str(row[0]).replace(',', '.')) for row in rows if
                row[0])

            conn.close()
            total_cost_label.config(text="ΣΥΝΟΛΟ ΕΥΡΩ : {:.2f} €".format(total_euros))

            return total_euros







        total_cost_label = Label(total_w, text=" ", font=("arial black", 13), width=30, background="#BAE5F1",
                                  relief="ridge", borderwidth=10)
        total_cost_label.grid(row=2, columnspan=2, pady=30, padx=60, ipadx=50)

        total_cost_btn = Button(total_w, text=" ΣΥΝΟΛΙΚΟ ΚΟΣΤΟΣ ΕΡΓΑΤΩΝ ", font=("arial black", 11), command=total_cost_workers, bd=6,
                           background="#06D784", activebackground="#06D784", activeforeground="#06D784")
        total_cost_btn.grid(row=1, pady=(40, 5), columnspan=2, padx=60, ipadx=35)

        title_label = Label(total_w, text="     ΕΡΓΑΤΕΣ     ", font=("arial black", 18), background="black",
                            relief="ridge", borderwidth=7, foreground="yellow")
        title_label.grid(row=0, columnspan=2, padx=210, pady=15)





    date_label = Label(root, text=" Ημερομηνία : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
    date_label.grid(row=1, column=0, sticky=W, padx=20, pady=(20, 10))
    name_label = Label(root, text=" Όνομα : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
    name_label.grid(row=2, column=0, sticky=W, padx=20, pady=10)
    work_label = Label(root, text=" Εργασία : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
    work_label.grid(row=3, column=0, sticky=W, padx=20, pady=10)
    hours_label = Label(root, text=" Ώρες : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
    hours_label.grid(row=4, column=0, sticky=W, padx=20, pady=10)
    price_label = Label(root, text=" Ημερομίσθιο € : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
    price_label.grid(row=5, column=0, sticky=W, padx=20, pady=10)
    choose_id_label = Label(root, text=" Επιλογή ID : ", font=("arial black", 12), background="#DBF0D7", foreground="#110464")
    choose_id_label.grid(row=6, column=0, sticky=W, padx=20, pady=10)

    date = Entry(root, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
    date.grid(row=1, column=1, padx=20, sticky=W, pady=(20, 10))
    name = Entry(root, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
    name.grid(row=2, column=1, padx=20, sticky=W, pady=10)
    work = Entry(root, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
    work.grid(row=3, column=1, padx=20, sticky=W, pady=10)
    hours = Entry(root, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
    hours.grid(row=4, column=1, padx=20, sticky=W, pady=10)
    price = Entry(root, width=20, font=("arial black", 12), bd=3, background="#DBE5E5")
    price.grid(row=5, column=1, padx=20, sticky=W, pady=10)
    choose_id = Entry(root, width=7, font=("arial black", 12), bd=3, background="#DBE5E5")
    choose_id.grid(row=6, column=1, padx=20, sticky=W, pady=10)

    worker_names = []


    submit_btn = Button(root, text="Εισαγωγή Στοιχείων Εργάτη", command=submit, font=("arial black", 11), bd=4, background="#06D784", activebackground="#06D784", activeforeground="#06D784")
    submit_btn.grid(row=8, column=0, pady=(20, 10), padx=10, ipadx=30)

    show_btn = Button(root, text="Εμφάνιση Στοιχείων Εργάτη", command=show, font=("arial black", 11), bd=4,  background="#0499A0", activebackground="#0499A0", activeforeground="#0499A0")
    show_btn.grid(row=9, column=1,  pady=10, padx=10, ipadx=38)

    delete_btn = Button(root, text="Διαγραφή Στοιχείων Εργάτη", command=delete, font=("arial black", 11), bd=4, background="#F67EA3", activebackground="#F67EA3", activeforeground="#F67EA3")
    delete_btn.grid(row=9, column=0, pady=10, padx=10, ipadx=30)

    edit_btn = Button(root, text="Επεξεργασία Στοιχείων Εργάτη", command=edit, font=("arial black", 11), bd=4, background="#A481C9", activebackground="#A481C9", activeforeground="#A481C9")
    edit_btn.grid(row=8, column=1, pady=(20, 10), padx=10, ipadx=28)

    select_btn = Button(root, text="  Επιλογή Εργάτη : ", command=select_worker, font=("arial black", 11), background="#0FAFC3")
    select_btn.grid(row=7, column=0, sticky=W, padx=10, pady=10)

    total_workers_cost_btn = Button(root, text="  ΣΥΝΟΛΙΚΟ ΚΟΣΤΟΣ ΕΡΓΑΤΩΝ : ", command=open_cost_worker, font=("arial black", 12), bd=6, background="#0FAFC3")
    total_workers_cost_btn.grid(row=10, columnspan=2, padx=10, pady=10)

    name_label = Label(root, text="Created and Designed by : Papaioannou Antonios", font=("arial black", 11), foreground="grey", background="#DBF0D7", borderwidth=1)
    name_label.grid(columnspan=2, row=11, sticky=E, pady=(5, 0))





    conn.commit()


    root.mainloop()